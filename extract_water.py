#!/usr/bin/env python3
"""提取水单明细数据 — 时段/渠道/价格带/技师人效/翻台率"""
import openpyxl, json, sys, os
from collections import Counter, defaultdict
from datetime import datetime

def extract_water_bill(report_path):
    wb = openpyxl.load_workbook(report_path, data_only=True)
    ws = None
    for s in wb.sheetnames:
        if '水单' in s: ws = wb[s]; break
    if not ws:
        print("No 水单 sheet found")
        return None
    
    hourly = Counter()
    dow = Counter()
    bands = Counter()
    channels = Counter()
    channel_rev = defaultdict(float)
    channel_gap = defaultdict(float)
    tech_orders = Counter()
    tech_rev = defaultdict(float)
    total_orders = 0
    total_received = 0
    total_discount = 0
    
    for r in range(2, ws.max_row + 1):
        ot = str(ws.cell(r, 2).value or '')
        if '服务' not in ot: continue
        
        t = ws.cell(r, 7).value
        rev_orig = float(ws.cell(r, 8).value or 0)
        rev_recv = float(ws.cell(r, 9).value or 0)
        pay = str(ws.cell(r, 11).value or '')
        tech = str(ws.cell(r, 13).value or '')
        tech_perf = float(ws.cell(r, 14).value or 0)
        
        if not t: continue
        if isinstance(t, datetime): h, d = t.hour, t.weekday()
        elif isinstance(t, str):
            try: dt = datetime.strptime(t[:19], '%Y-%m-%d %H:%M:%S'); h, d = dt.hour, dt.weekday()
            except: continue
        else: continue
        
        hourly[h] += 1; dow[d] += 1; total_orders += 1
        total_received += rev_recv; total_discount += (rev_orig - rev_recv)
        tech_orders[tech] += 1; tech_rev[tech] += rev_orig
        
        if rev_recv < 80: bands['<¥80'] += 1
        elif rev_recv < 120: bands['¥80-120'] += 1
        elif rev_recv < 160: bands['¥120-160'] += 1
        elif rev_recv < 200: bands['¥160-200'] += 1
        elif rev_recv < 260: bands['¥200-260'] += 1
        else: bands['¥260+'] += 1
        
        if '美团' in pay: k = '美团'
        elif '抖音' in pay: k = '抖音'
        elif any(x in pay for x in ['现金','微信','支付宝','扫','转账']): k = '直付'
        else: continue
        channels[k] += 1; channel_rev[k] += rev_recv; channel_gap[k] += (rev_orig - rev_recv)
    
    wb.close()
    
    # Compute metrics
    import calendar
    year, month = int(report_path.split('年')[1][:4]) if '年' in report_path else 2026, int(report_path.split('月')[0][-2:]) if '月' in report_path else 6
    from datetime import date
    # Parse year/month from filename or use defaults
    try:
        parts = report_path.replace('.xlsx','').split('年')
        if len(parts) > 1:
            yr = int(parts[0][-4:])
            mo = int(parts[1].split('月')[0])
        else:
            yr, mo = 2026, 6
    except:
        yr, mo = 2026, 6
    days_in_month = calendar.monthrange(yr, mo)[1]
    beds = 17
    daily = total_orders / days_in_month
    wd = sum(dow[d] for d in range(5))
    we = sum(dow[d] for d in range(5, 7))
    wd_days = len([d for d in range(1, days_in_month+1) if date(yr, mo, d).weekday() < 5])
    we_days = days_in_month - wd_days
    
    result = {
        'total_orders': total_orders,
        'total_received': round(total_received, 2),
        'total_discount': round(total_discount, 2),
        'daily_avg': round(daily, 1),
        'per_bed': round(daily / beds, 2),
        'wkday_avg': round(wd / wd_days) if wd_days else 0,
        'wkend_avg': round(we / we_days) if we_days else 0,
        'hourly_pct': [round(hourly.get(h, 0) / total_orders * 100, 1) if total_orders else 0 for h in range(24)],
        'day_orders': [dow[d] for d in range(7)],
        'price_bands': {b: {'orders': bands[b], 'pct': round(bands[b]/total_orders*100,1)} for b in ['<¥80','¥80-120','¥120-160','¥160-200','¥200-260','¥260+']},
        'channels': {ch: {'orders': channels[ch], 'pct': round(channels[ch]/total_orders*100,1), 'avg_gap': round(channel_gap[ch]/channels[ch])} for ch in ['美团','抖音','直付']},
        'technicians': [{'name': t, 'orders': tech_orders[t], 'daily': round(tech_orders[t]/30,1), 'revenue': round(tech_rev[t])} for t in sorted(tech_orders, key=lambda x: tech_orders[x], reverse=True)],
    }
    
    return result

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 extract_water.py <report.xlsx>")
        sys.exit(1)
    
    data = extract_water_bill(sys.argv[1])
    if data:
        print(json.dumps(data, indent=2, ensure_ascii=False))
        # Save to project
        base = os.path.dirname(os.path.abspath(__file__))
        out_path = os.path.join(base, '02_经营分析', '水单明细数据.json')
        with open(out_path, 'w') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"\nSaved to {out_path}")
