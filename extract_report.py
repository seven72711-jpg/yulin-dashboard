#!/usr/bin/env python3
"""
十三绝门店月度报表数据提取器
输入：月度报表 Excel 文件路径
输出：结构化 JSON，包含所有经营分析需要的核心数据
用法：python extract_report.py /path/to/report.xlsx
"""

import openpyxl
import json
import sys
import os
from datetime import datetime
from collections import defaultdict


def safe_float(v):
    try:
        return float(v) if v is not None else 0
    except (ValueError, TypeError):
        return 0


def extract_all(filepath):
    """从月报中提取所有核心数据"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {
        "source": os.path.basename(filepath),
        "extracted_at": datetime.now().isoformat(),
    }

    # ========== 1. 店面经营报表 ==========
    ws = wb['店面经营报表']
    
    # Detect year ranges from merged headers
    # Columns: 2024=C(3)-P(16), 2025=Q(17)-AD(30), 2026=AF(32)-AS(45)
    layout = {
        '2024': {'start': 3, 'end': 16, 'months': ['01','02','03','04','05','06','07','08','09','10','11','12'], 'cum_col': 15, 'pct_col': 16},
        '2025': {'start': 17, 'end': 30, 'months': ['01','02','03','04','05','06','07','08','09','10','11','12'], 'cum_col': 29, 'pct_col': 30},
        '2026': {'start': 32, 'end': 45, 'months': ['01','02','03','04','05','06','07','08','09','10','11','12'], 'cum_col': 44, 'pct_col': 45},
    }

    monthly = []
    
    for year, cfg in layout.items():
        for i, month in enumerate(cfg['months']):
            col = cfg['start'] + i
            if col > cfg['end']:
                break
            
            revenue = safe_float(ws.cell(row=3, column=col).value)  # 主营收入
            if revenue == 0:
                continue
            
            other_income = safe_float(ws.cell(row=4, column=col).value)
            total_revenue = safe_float(ws.cell(row=5, column=col).value)
            
            entry = {
                "period": f"{year}-{month}",
                "revenue": round(total_revenue, 2),
                "main_revenue": round(revenue, 2),
                "other_income": round(other_income, 2),
                
                # 人力成本
                "labor_salary": round(safe_float(ws.cell(row=6, column=col).value), 2),
                "labor_social": round(safe_float(ws.cell(row=7, column=col).value), 2),
                "labor_meals": round(safe_float(ws.cell(row=8, column=col).value), 2),
                "labor_dorm_util": round(safe_float(ws.cell(row=9, column=col).value), 2),
                "labor_dorm_rent": round(safe_float(ws.cell(row=10, column=col).value), 2),
                "labor_benefits": round(safe_float(ws.cell(row=11, column=col).value), 2),
                "labor_training": round(safe_float(ws.cell(row=12, column=col).value), 2),
                "labor_uniform": round(safe_float(ws.cell(row=13, column=col).value), 2),
                "labor_total": round(safe_float(ws.cell(row=14, column=col).value), 2),
                
                # 店铺基本成本
                "mgmt_fee": round(safe_float(ws.cell(row=15, column=col).value), 2),
                "rent": round(safe_float(ws.cell(row=16, column=col).value), 2),
                "utilities": round(safe_float(ws.cell(row=17, column=col).value), 2),
                "property": round(safe_float(ws.cell(row=18, column=col).value), 2),
                "office": round(safe_float(ws.cell(row=19, column=col).value), 2),
                "bank_fees": round(safe_float(ws.cell(row=20, column=col).value), 2),
                "consumables": round(safe_float(ws.cell(row=21, column=col).value), 2),
                "non_consumables": round(safe_float(ws.cell(row=22, column=col).value), 2),
                "repairs": round(safe_float(ws.cell(row=23, column=col).value), 2),
                "travel": round(safe_float(ws.cell(row=24, column=col).value), 2),
                "phone": round(safe_float(ws.cell(row=25, column=col).value), 2),
                "refunds": round(safe_float(ws.cell(row=26, column=col).value), 2),
                "accidents": round(safe_float(ws.cell(row=27, column=col).value), 2),
                "interest": round(safe_float(ws.cell(row=28, column=col).value), 2),
                "tax_fees": round(safe_float(ws.cell(row=29, column=col).value), 2),
                "base_cost_total": round(safe_float(ws.cell(row=30, column=col).value), 2),
                
                # 营销成本
                "promotion": round(safe_float(ws.cell(row=31, column=col).value), 2),
                "advertising": round(safe_float(ws.cell(row=32, column=col).value), 2),
                "free_service_cost": round(safe_float(ws.cell(row=33, column=col).value), 2),
                "koc_visits": round(safe_float(ws.cell(row=34, column=col).value), 2),
                "marketing_total": round(safe_float(ws.cell(row=35, column=col).value), 2),
                
                # 汇总
                "total_cost": round(safe_float(ws.cell(row=36, column=col).value), 2),
                "gross_profit": round(safe_float(ws.cell(row=37, column=col).value), 2),
                "manager_bonus": round(safe_float(ws.cell(row=38, column=col).value), 2),
                "operating_profit": round(safe_float(ws.cell(row=39, column=col).value), 2),
                
                # 税务行（行42-43可能有不同年份的差异）
                "income_tax": round(safe_float(ws.cell(row=42, column=col).value), 2),
                "net_profit": round(safe_float(ws.cell(row=43, column=col).value), 2),
            }
            
            # 计算比率
            if entry["revenue"] > 0:
                entry["labor_ratio"] = round(entry["labor_total"] / entry["revenue"] * 100, 1)
                entry["net_margin"] = round(entry["net_profit"] / entry["revenue"] * 100, 1)
                entry["base_cost_ratio"] = round(entry["base_cost_total"] / entry["revenue"] * 100, 1)
                entry["marketing_ratio"] = round(entry["marketing_total"] / entry["revenue"] * 100, 1)
            
            monthly.append(entry)
    
    result["monthly"] = monthly

    # ========== 2. 水单明细 → 渠道 & 储值分析 ==========
    # Sheet name varies: "十三绝系统" vs "伊智系统"
    detail_name = None
    for s in wb.sheetnames:
        if '水单明细' in s:
            detail_name = s
            break
    if not detail_name:
        raise KeyError("找不到水单明细 sheet")
    ws_detail = wb[detail_name]
    
    channels = defaultdict(lambda: {"orders": 0, "raw": 0, "received": 0})
    order_types = defaultdict(lambda: {"count": 0, "amount": 0})
    stored_value = {"充值": 0, "开卡": 0, "疗程套餐": 0, "卡金消费": 0}
    total_service_perf = 0
    
    # Detect format: 伊智 (old, ~25 cols) vs 十三绝 (new, ~33 cols)
    is_new_system = ws_detail.max_column > 30
    
    for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row, max_col=ws_detail.max_column, values_only=True):
        try:
            order_type = str(row[1] or '').strip()  # col 2 = index 1
            
            if is_new_system:
                # === 十三绝系统: 33 columns ===
                raw_price = safe_float(row[7])     # col 8
                received = safe_float(row[8])       # col 9
                perf = safe_float(row[9])           # col 10
                platform = str(row[21] or '').strip()  # col 22
                card_perf = safe_float(row[29])     # col 30
                cash_perf = safe_float(row[27])     # col 28
                other_perf = safe_float(row[31])    # col 32
            else:
                # === 伊智系统: 25 columns ===
                raw_price = safe_float(row[8])      # col 9 = 门店业绩
                received = safe_float(row[8])       # col 9 (same, no separate received)
                perf = safe_float(row[8])           # col 9
                platform = ''                       # no platform column
                card_perf = safe_float(row[21])     # col 22
                cash_perf = safe_float(row[19])     # col 20
                other_perf = safe_float(row[23])    # col 24
                # Normalize: "开单" = service
                if '开单' in order_type:
                    order_type = '服务项目'
            if raw_price == 0:
                continue
            
            channels[platform]['orders'] += 1
            channels[platform]['raw'] += raw_price
            channels[platform]['received'] += received
            
            order_types[order_type]['count'] += 1
            order_types[order_type]['amount'] += received
            
            if '充值' in order_type:
                stored_value['充值'] += received
            elif '开卡' in order_type:
                stored_value['开卡'] += received
            elif '疗程' in order_type:
                stored_value['疗程套餐'] += received
            else:
                total_service_perf += raw_price
                stored_value['卡金消费'] += card_perf
                
        except Exception:
            continue
    
    # Dedup: 卡金消费 already counted
    channel_summary = {}
    for ch, data in channels.items():
        ch_name = ch if ch else '直付/现金'
        channel_summary[ch_name] = {
            "orders": data['orders'],
            "raw": round(data['raw'], 2),
            "received": round(data['received'], 2),
            "discount_rate": round((1 - data['received']/data['raw'])*100, 1) if data['raw'] > 0 else 0,
            "share": round(data['received']/sum(c['received'] for c in channels.values())*100, 1) if sum(c['received'] for c in channels.values()) > 0 else 0,
        }
    
    result["water_bill"] = {
        "channels": channel_summary,
        "order_types": {k: {"count": v['count'], "amount": round(v['amount'], 2)} for k, v in sorted(order_types.items())},
        "stored_value": {
            "inflow_充值": round(stored_value['充值'], 2),
            "inflow_开卡": round(stored_value['开卡'], 2),
            "inflow_套餐": round(stored_value['疗程套餐'], 2),
            "inflow_total": round(stored_value['充值'] + stored_value['开卡'] + stored_value['疗程套餐'], 2),
            "outflow_消费": round(stored_value['卡金消费'], 2),
            "net_change": round(stored_value['充值'] + stored_value['开卡'] + stored_value['疗程套餐'] - stored_value['卡金消费'], 2),
        },
        "service_revenue": round(total_service_perf, 2),
    }

    # ========== 3. 店长提成 ==========
    try:
        ws_bonus = wb['店长提成奖励']
        result["manager_bonus"] = {
            "actual_performance": round(safe_float(ws_bonus.cell(row=1, column=2).value), 2),
            "target_performance": round(safe_float(ws_bonus.cell(row=1, column=4).value), 2),
            "salary_cost": round(safe_float(ws_bonus.cell(row=2, column=2).value), 2),
            "utilities_cost": round(safe_float(ws_bonus.cell(row=3, column=2).value), 2),
            "consumables_cost": round(safe_float(ws_bonus.cell(row=5, column=2).value), 2),
            "total_manager_cost": round(safe_float(ws_bonus.cell(row=6, column=2).value), 2),
            "manager_margin": round(safe_float(ws_bonus.cell(row=7, column=2).value), 2),
            "bonus_amount": round(safe_float(ws_bonus.cell(row=14, column=2).value), 2),
        }
    except:
        result["manager_bonus"] = None

    # ========== 4. 总部结算 ==========
    try:
        ws_settle = wb[[s for s in wb.sheetnames if '总部结算' in s][0]]
        deductions = []
        settle_amount = 0
        for row in ws_settle.iter_rows(min_row=3, max_row=ws_settle.max_row, max_col=7, values_only=True):
            desc = str(row[1] or '').strip()
            amount = safe_float(row[2])
            if not desc:
                continue
            if '结算金额' in desc or '应结算' in desc:
                settle_amount = amount
            else:
                deductions.append({"item": desc, "amount": round(amount, 2)})
        
        result["settlement"] = {
            "deductions": deductions,
            "total_deductions": round(sum(d['amount'] for d in deductions), 2),
            "net_settlement": round(settle_amount, 2),
        }
    except:
        result["settlement"] = None

    # ========== 5. 银行流水摘要 ==========
    try:
        bank_name = None
        for s in wb.sheetnames:
            if '银行交易流水' in s:
                bank_name = s
                break
        if bank_name:
            ws_bank = wb[bank_name]
        cash_flows = {"inflow": [], "outflow": []}
        for row in ws_bank.iter_rows(min_row=4, max_row=ws_bank.max_row, max_col=11, values_only=True):
            date_val = str(row[0] or '')
            purpose = str(row[1] or '').strip()
            debit = safe_float(row[2])  # 支出
            credit = safe_float(row[3])  # 收入
            counterparty = str(row[6] or '').strip()
            
            if debit > 0:
                cash_flows['outflow'].append({
                    "date": date_val[:10] if date_val else '',
                    "purpose": purpose,
                    "amount": round(debit, 2),
                    "counterparty": counterparty,
                })
            elif credit > 0:
                cash_flows['inflow'].append({
                    "date": date_val[:10] if date_val else '',
                    "purpose": purpose,
                    "amount": round(credit, 2),
                    "counterparty": counterparty,
                })
        
        result["bank_flows"] = cash_flows
    except:
        result["bank_flows"] = None

    # ========== 6. 核心指标一键速览 ==========
    if monthly:
        latest = monthly[-1]
        result["dashboard"] = {
            "period": latest["period"],
            "revenue": latest["revenue"],
            "net_profit": latest["net_profit"],
            "net_margin": latest["net_margin"],
            "labor_ratio": latest["labor_ratio"],
            "marketing_ratio": latest["marketing_ratio"],
            "payroll": latest["labor_salary"],
            "rent": latest["rent"],
            "promotion_fee": latest["promotion"],
            "mgmt_fee": latest["mgmt_fee"],
        }
        
        # 卡金数据
        if "water_bill" in result:
            sv = result["water_bill"]["stored_value"]
            result["dashboard"]["stored_value_in"] = sv["inflow_total"]
            result["dashboard"]["stored_value_out"] = sv["outflow_消费"]
            result["dashboard"]["stored_value_net"] = sv["net_change"]
    
    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_report.py <report.xlsx>")
        sys.exit(1)
    
    data = extract_all(sys.argv[1])
    print(json.dumps(data, ensure_ascii=False, indent=2))
